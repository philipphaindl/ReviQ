import os
from tabulate import tabulate
from modules.common import *
from modules.bibtex_processing import *
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from datetime import date

headers = ["Library", "Retrieved", "Duplicates", "Automatically Excluded", "Manual Evaluation", "Resulting", "Precision %", "Recall %", "F-Score %"]

def process_statistics(wb, it_config, iteration_name, metrics, total_selected, selection_result):
    """Process and document the statistics of the paper selection process"""
    create_statistics_sheet(wb)
    summarize_selection_criteria(it_config, iteration_name, metrics, total_selected, selection_result)
    fill_statistics_sheet(wb, iteration_name, metrics)

def create_statistics_sheet(wb):
    """Create the Excel sheet reporting the numbers of excluded and included papers based on the criteria"""
    if STATISTICS_SHEET_NAME in wb.sheetnames:
        statistics_sheet = wb[STATISTICS_SHEET_NAME]
    else:
        statistics_sheet = wb.create_sheet(STATISTICS_SHEET_NAME,1)

    statistics_sheet.cell(row=1, column=1).value = "Date"
    statistics_sheet.cell(row=1, column=1).font = Font(bold=True)
    statistics_sheet.column_dimensions["A"].width = 15

    statistics_sheet.cell(row=1, column=2).value = "Iteration"
    statistics_sheet.cell(row=1, column=2).font = Font(bold=True)
    statistics_sheet.column_dimensions["B"].width = 20

    statistics_sheet.cell(row=1, column=3).value = headers[0]
    statistics_sheet.cell(row=1, column=3).font = Font(bold=True)
    statistics_sheet.column_dimensions["C"].width = 15

    statistics_sheet.cell(row=1, column=4).value = headers[1]
    statistics_sheet.cell(row=1, column=4).font = Font(bold=True)
    statistics_sheet.column_dimensions["D"].width = 10

    statistics_sheet.cell(row=1, column=5).value = headers[2]
    statistics_sheet.cell(row=1, column=5).font = Font(bold=True)
    statistics_sheet.column_dimensions["E"].width = 10

    for n in range(1, 4):
        ec = "ec" + str(n)
        ec = ec.upper()
        statistics_sheet.cell(row=1, column=n+5).value = ec
        statistics_sheet.cell(row=1, column=n+5).font = Font(bold=True)
        statistics_sheet.column_dimensions[get_column_letter(n+5)].width = 8
        statistics_sheet.cell(row=1, column=n+5).alignment = Alignment(wrapText=True,horizontal='center')

    statistics_sheet.cell(row=1, column=9).value = headers[4]
    statistics_sheet.cell(row=1, column=9).font = Font(bold=True)
    statistics_sheet.column_dimensions["I"].width = 12
    statistics_sheet.cell(row=1, column=9).alignment = Alignment(wrapText=True,horizontal='center')

    for n in range(4, 11):
        ec = "ec" + str(n)
        ec = ec.upper()
        statistics_sheet.cell(row=1, column=n+6).value = ec
        statistics_sheet.cell(row=1, column=n+6).font = Font(bold=True)
        statistics_sheet.column_dimensions[get_column_letter(n+6)].width = 8
        statistics_sheet.cell(row=1, column=n+6).alignment = Alignment(wrapText=True,horizontal='center')

    for n in range(1, 11):
        ic = "ic" + str(n)
        ic = ic.upper()
        statistics_sheet.cell(row=1, column=n+16).value = ic
        statistics_sheet.cell(row=1, column=n+16).font = Font(bold=True)
        statistics_sheet.column_dimensions[get_column_letter(n+16)].width = 8
        statistics_sheet.cell(row=1, column=n+16).alignment = Alignment(wrapText=True,horizontal='center')

    for n in range(1, 11):
        qc = "qc" + str(n)
        qc = qc.upper()
        statistics_sheet.cell(row=1, column=n+26).value = qc
        statistics_sheet.cell(row=1, column=n+26).font = Font(bold=True)
        statistics_sheet.column_dimensions[get_column_letter(n+26)].width = 8
        statistics_sheet.cell(row=1, column=n+26).alignment = Alignment(wrapText=True,horizontal='center')

    statistics_sheet.cell(row=1, column=37).value = headers[5]
    statistics_sheet.cell(row=1, column=37).font = Font(bold=True)
    statistics_sheet.column_dimensions[get_column_letter(37)].width = 12
    statistics_sheet.cell(row=1, column=37).alignment = Alignment(wrapText=True,horizontal='center')

    statistics_sheet.cell(row=1, column=38).value = headers[6]
    statistics_sheet.cell(row=1, column=38).font = Font(bold=True)
    statistics_sheet.column_dimensions[get_column_letter(38)].width = 12
    statistics_sheet.cell(row=1, column=38).alignment = Alignment(wrapText=True,horizontal='center')

    statistics_sheet.cell(row=1, column=39).value = headers[7]
    statistics_sheet.cell(row=1, column=39).font = Font(bold=True)
    statistics_sheet.column_dimensions[get_column_letter(39)].width = 12
    statistics_sheet.cell(row=1, column=39).alignment = Alignment(wrapText=True,horizontal='center')

    statistics_sheet.cell(row=1, column=40).value = headers[8]
    statistics_sheet.cell(row=1, column=40).font = Font(bold=True)
    statistics_sheet.column_dimensions[get_column_letter(40)].width = 12
    statistics_sheet.cell(row=1, column=40).alignment = Alignment(wrapText=True,horizontal='center')

def fill_statistics_sheet(wb, iteration_name, metrics):
    """Fill the statistics sheet""" 
    sheet = wb[STATISTICS_SHEET_NAME]
    row_id = find_empty_row(sheet)
    
    output_entries = {header: [] for header in headers}
    
    for library, data in metrics.items():
        sheet.cell(row=row_id, column=1, value=date.today()).alignment = Alignment(horizontal='left')
        sheet.cell(row=row_id, column=2, value=iteration_name)
        sheet.cell(row=row_id, column=3, value=library[:-4])
        output_entries["Library"].append(library[:-4])
        
        automatically_excluded = sum(data[i] for i in [2, 3, 4])
        output_entries["Automatically Excluded"].append(automatically_excluded)
        
        for i, metric in enumerate(data):
            sheet.cell(row=row_id, column=4+i, value=metric).alignment = Alignment(horizontal='center')
            write_selections_per_step(output_entries, i, metric)
        
        row_id += 1
    
    print(tabulate(output_entries, headers=headers))

def find_empty_row(sheet):
    """Finds the first empty row in a sheet"""
    row_id = 1
    while sheet.cell(row=row_id, column=1).value:
        row_id += 1
    return row_id

def summarize_selection_criteria(it_config, iteration_name, metrics, total_selected, selection_result):
    """Summarize the selection process mandated by the inclusion, exclusion, and quality criteria and write it to Excel"""
    final_entries = []
    retrieved_dir = os.path.join(it_config.iteration_dir, RETRIEVED_DIRNAME)
    
    for filename in filter(lambda f: f.endswith('.bib'), os.listdir(retrieved_dir)):
        resulting_entries = get_bibtex_entries(it_config.iteration_dir, AFTER_MANUAL_EVALUATION_DIRNAME, filename)
        final_entries.extend(resulting_entries)
        
        library = filename[:-4]
        papers_of_library = selection_result.papers[library]
        
        aggregate_selection_criteria(metrics, filename, selection_result, papers_of_library, prefix='ec', start=4, end=11)
        aggregate_selection_criteria(metrics, filename, selection_result, papers_of_library, prefix='ic', start=1, end=11)
        aggregate_selection_criteria(metrics, filename, selection_result, papers_of_library, prefix='qc', start=1, end=11)
        
        save_performance_metrics(metrics, filename, resulting_entries, total_selected)
    
def aggregate_selection_criteria(metrics, filename, selection_result, papers, prefix, start, end):
    """Aggregate the selection criteria that were assigned to a paper"""
    for n in range(start, end):
        criteria_key = f"{prefix}{n}"
        count = sum(1 for paper in papers if criteria_key in selection_result.results and paper.citekey in selection_result.results[criteria_key])
        metrics[filename].append(count)

def calculate_performance_metrics(retrieved, selected, total_selected):
    """Calculate precision, recall and f-score for a set of retrieved vs. selected papers"""
    if retrieved == 0 or total_selected == 0:
        return "0.00 (0/0)", "0.00 (0/0)", "0.00"
    
    precision = selected / retrieved if retrieved > 0 else 0
    recall = selected / total_selected if total_selected > 0 else 0
    
    f_score = (200 * (precision * recall) / (precision + recall)) if (precision + recall) > 0 else 0
    
    precision_txt = f"{round(100 * precision, 2)} ({selected}/{retrieved})"
    recall_txt = f"{round(100 * recall, 2)} ({selected}/{total_selected})"
    f_score_txt = f"{round(f_score, 2)}"
    
    return precision_txt, recall_txt, f_score_txt

def write_selections_per_step(output_entries, index, metric):
    """Writes the number of selected papers per step"""
    key_map = {
        0: "Retrieved", 1: "Duplicates", 5: "Manual Evaluation",
        33: "Resulting", 34: "Precision %", 35: "Recall %", 36: "F-Score %"
    }
    if index in key_map:
        output_entries[key_map[index]].append(metric)

def save_performance_metrics(metrics, filename, resulting_entries, total_selected):
    """Save the calculated performance metrics for a particular library"""
    selected = len(resulting_entries)
    metrics[filename].append(selected)
    retrieved = metrics[filename][0]
    precision, recall, f_score = calculate_performance_metrics(retrieved, selected, total_selected)
    metrics[filename].extend([precision, recall, f_score])