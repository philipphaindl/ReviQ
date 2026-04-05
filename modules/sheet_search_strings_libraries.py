from modules.common import *
from openpyxl.styles import Font, Alignment
from datetime import date

def process_search_strings_and_libraries(wb, config):
    """Process and document the libraries and search metadata of the paper selection process"""
    create_search_strings_and_libraries_sheet(wb, config)
    fill_search_strings_and_libraries_sheet(wb, config)

def create_search_strings_and_libraries_sheet(wb, config):
    """Create Excel sheet documenting the search strings, libraries, and applied filters"""
    if not wb:
        raise ValueError("Workbook object cannot be None")
    
    if LIBRARIES_SHEET_NAME in wb.sheetnames:
        sheet = wb[LIBRARIES_SHEET_NAME]
    else:
        sheet = wb.active
        sheet.title = LIBRARIES_SHEET_NAME
    
    headers = [
        ("Date", 15),
        ("Iteration", 20),
        ("Library", 10),
        ("Search String", 70),
        ("Filters", 80),
        ("Quality Score Threshold", 15, 'center', True)
    ]
    
    for col_idx, (header, width, *alignment) in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        sheet.column_dimensions[cell.column_letter].width = width
        
        if alignment:
            cell.alignment = Alignment(horizontal=alignment[0], wrapText=True if len(alignment) > 1 else False)
    
    if "iterations" not in config:
        raise KeyError("Config is missing 'iterations' key")
    
def fill_search_strings_and_libraries_sheet(wb, config):
    """Document name, query, filter, and threshold for each library""" 
    sheet = wb["Libraries and Search Strings"]
    row = 2
    for iteration_config in config["iterations"]:
        threshold = iteration_config["quality-score-treshold"]
        if threshold is None:
            raise KeyError("Missing 'quality-score-treshold' in iteration config")
        
        for library in iteration_config.get("libraries", []):
            sheet.append([
                date.today(),
                iteration_config.get("dir", ""),
                library.get("name", ""),
                library.get("query", ""),
                library.get("filter", ""),
                threshold
            ])
            
            sheet.cell(row=row, column=5).alignment = Alignment(wrapText=True)
            sheet.cell(row=row, column=6).alignment = Alignment(horizontal='center')
            row += 1
