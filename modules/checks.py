import os
import openpyxl
from modules.common import *

def check_manual_evaluations(it_conf):
    """Validates the manual check Excel file for missing values and conflicting criteria."""
    excel_file_with_manual_evaluations = os.path.join(it_conf.iteration_dir, AFTER_MANUAL_EVALUATION_DIRNAME, MANUAL_EVALUATION_FILENAME)
    if not os.path.exists(excel_file_with_manual_evaluations):
        print(f"Error: Need file {excel_file_with_manual_evaluations}")
        exit(1)
    
    wb = openpyxl.load_workbook(excel_file_with_manual_evaluations)
    manual_check_sheet = wb.active
    validate_study_attributes(manual_check_sheet, excel_file_with_manual_evaluations)

def validate_study_attributes(sheet, excel_file_with_manual_evaluations):
    """Checks for missing values and conflicting inclusion/exclusion criteria."""
    row_id = 2
    
    while True:
        title_cell = sheet.cell(row=row_id, column=1).value
        if not title_cell:
            break

        validate_missing_values(sheet, row_id, excel_file_with_manual_evaluations)
        validate_conflicting_criteria(sheet, row_id, excel_file_with_manual_evaluations)
                
        row_id += 1

def validate_missing_values(sheet, row_id, excel_file_with_manual_evaluations):
    """Ensures that required fields are filled unless explicitly excluded."""
    exclusion_criterion = sheet.cell(row=row_id, column=4).value
    if not (exclusion_criterion and exclusion_criterion.startswith("EC")):
        """ User has not excluded this study, so there must be inclusion and quality criteria """ 

        inclusion_criterion = sheet.cell(row=row_id, column=6).value if sheet.cell(row=row_id, column=6) else None
        quality_criterion = sheet.cell(row=row_id, column=5).value if sheet.cell(row=row_id, column=5) else None

        if not (inclusion_criterion and quality_criterion):
            print(f"Line {row_id}: Missing values detected in file {excel_file_with_manual_evaluations}")
            exit(1)

def validate_conflicting_criteria(sheet, row_id, excel_file_with_manual_evaluations):
    """Checks that a study is not associated with both inclusion and exclusion criteria."""
    if sheet.cell(row=row_id, column=4).value and sheet.cell(row=row_id, column=6).value:
        print(f"Line {row_id}: A study in {excel_file_with_manual_evaluations} has both inclusion and exclusion criteria")
        exit(1)