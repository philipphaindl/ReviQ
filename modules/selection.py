from modules.bibtex_processing import *
from datetime import date
import langid
from pathlib import Path
from openpyxl.styles import Font, Alignment

def eval_e1(citekey, title, abstract, keywords, entry_type, excluded, included, it_conf, success_clb, selection_result):
    """Exclusion Criterion 1: Title or abstract contain the relevant keywords"""
    title_match = check_text_contains_keywords(title, it_conf)
    abstract_match = check_text_contains_keywords(abstract, it_conf)
    keywords_match = check_text_contains_keywords(keywords, it_conf)

    if not(title_match or abstract_match or keywords_match):
        excluded.append(citekey)
        success_clb("ec1", citekey, selection_result.results)
    else:
        included.append(citekey)

def check_text_contains_keywords(text, it_conf):
    """Check if the given text contains the relevant keywords"""
    if not text:
        return False
    
    for keyword in it_conf.keywords:
        if keyword in text.lower():
            return True

    return False

def eval_e2(citekey, title, abstract, keywords, entry_type, excluded, included, it_conf, success_clb, selection_result):
    """Exclusion Criterion 2: Language of paper is English"""
    if citekey in included and abstract:
        language, _ = langid.classify(abstract)
        if language != 'en':
            included.remove(citekey)
            excluded.append(citekey)
            success_clb("ec2", citekey, selection_result.results)

def eval_e3(citekey, title, abstract, keywords, entry_type, excluded, included, it_conf, success_clb, selection_result):
    """Exclusion Criterion 3: Check if a paper is peer-reviewed, i.e., a conference or journal publication"""
    if citekey in included and entry_type:
        if entry_type.lower() not in ["inproceedings", "article", "conference"]:
            included.remove(citekey)
            excluded.append(citekey)
            success_clb("ec3", citekey, selection_result.results)

def evaluate_studies(it_config, sheet, row_id, duplicates_detection_config, selection_result):
    """Evaluate the studies of the given iteration w.r.t. exclusion criteria and duplicates"""
    directory = os.path.join(it_config.iteration_dir, RETRIEVED_DIRNAME)
    included_count = 0
    metrics = {}
    evalfuncs = [eval_e1, eval_e2, eval_e3]

    for filename in os.listdir(directory):
        if filename.endswith('.bib'):
            metrics[filename] = []
            included = []
            excluded = []
            file_path = os.path.join(directory, directory, filename)
            entries = read_bib_file(file_path)
            included_count = len(entries)
            metrics[filename].append(included_count)

            for entry in entries:
                document_paper(sheet, row_id, entry, filename, Path(it_config.iteration_dir).stem)
                library = filename[:-4]

                if not library in selection_result.papers:
                    selection_result.papers[library] = []

                author = entry.get("author")
                title = entry.get("title")
                year = entry.get("year")
                citekey = entry.get("ID")
                paper = Paper(author, title, year, citekey)
                selection_result.papers[library].append(paper)

                row_id = row_id + 1

            remove_duplicates(it_config.iteration_config, entries, os.path.join(it_config.iteration_dir, DEDUPLICATED_DIRNAME, filename), duplicates_detection_config, selection_result)
            deduplicated_entries = read_bib_file(os.path.join(it_config.iteration_dir, DEDUPLICATED_DIRNAME, filename))

            metrics[filename].append(included_count - len(deduplicated_entries))

            entries = deduplicated_entries
            included_count = len(deduplicated_entries)

            previous_excluded = 0
            for evalfunc in evalfuncs:
                evaluate_entries(entries, evalfunc, excluded, included, sheet, it_config, selection_result)
                diff = len(excluded) - previous_excluded
                previous_excluded = len(excluded)
                metrics[filename].append(diff)
                included_count -= diff

            included_entries = get_entries_by_citekey(entries, included)
            metrics[filename].append(included_count)

            output_path = os.path.join(it_config.iteration_dir, NEEDING_MANUAL_EVALUATION_DIRNAME, filename)
            write_bibtex_entries(output_path, included_entries)
     
    return row_id, metrics

def document_paper(sheet, row_id, entry, filename, iteration):
		sheet.cell(row=row_id, column=1).value = date.today()
		sheet.cell(row=row_id, column=1).alignment = Alignment(horizontal='left', vertical='top')

		sheet.cell(row=row_id, column=2).value = iteration
		sheet.cell(row=row_id, column=2).alignment = Alignment(vertical='top')

		library = filename[:-4]
		sheet.cell(row=row_id, column=3).value = library
		sheet.cell(row=row_id, column=3).alignment = Alignment(vertical='top')

		doi = entry.get("doi")
		sheet.cell(row=row_id, column=4).value = doi
		sheet.cell(row=row_id, column=4).alignment = Alignment(vertical='top')

		citekey = entry.get("ID")
		sheet.cell(row=row_id, column=5).value = citekey
		sheet.cell(row=row_id, column=5).alignment = Alignment(vertical='top')

		author = entry.get("author")
		sheet.cell(row=row_id, column=6).value = author
		sheet.cell(row=row_id, column=6).alignment = Alignment(wrapText=True,vertical='top')

		title = entry.get("title")
		sheet.cell(row=row_id, column=7).value = title
		sheet.cell(row=row_id, column=7).alignment = Alignment(wrapText=True,vertical='top')

		year = entry.get("year")
		sheet.cell(row=row_id, column=8).value = year
		sheet.cell(row=row_id, column=8).alignment = Alignment(horizontal='center',vertical='top')

		sheet.cell(row=row_id, column=9).value = "automatic"
		sheet.cell(row=row_id, column=9).alignment = Alignment(horizontal='center',vertical='top')

def evaluate_entries(entries, evalfunc, excluded, included, sheet, it_conf, selection_result):
    """Evaluate the given bibtex entries based on the first three exclusion criteria"""
    for entry in entries:
        title = entry.get('title', '')
        abstract = entry.get('abstract', '')
        keywords = entry.get('keywords', '')
        citekey = entry.get("ID", '')
        entry_type = entry.get("ENTRYTYPE", '')
        
        if not citekey:
            continue  # Skip entries without a valid citekey
        
        evalfunc(citekey, title, abstract, keywords, entry_type, excluded, included, it_conf, add_to_results, selection_result)
    
    return excluded, included

def is_known_duplicate(iteration_config, title):
    """Checks whether the given title is a known duplicate based on the configuration file"""
    if not iteration_config or "duplicates" not in iteration_config or not title:
        return False
    
    return any(known_duplicate.lower() == title.lower() for known_duplicate in iteration_config["duplicates"])

def remove_duplicates(iteration_config, entries, file, duplicates_detection_config, selection_result):
    """Remove duplicates from the given set of entries"""
    deduplicated_entries = []
    for entry in entries:
        if "doi" in entry:
            doi = entry["doi"]
            if not doi in duplicates_detection_config.dois:
                if "title" in entry:
                    title = entry["title"].lower()
                    title_and_venue = title + "_" + get_venue(entry).lower()
                    if not title_and_venue in duplicates_detection_config.titles_and_venues and not is_known_duplicate(iteration_config, title):
                        duplicates_detection_config.dois.append(doi)
                        duplicates_detection_config.titles_and_venues.append(title_and_venue.lower())
                        deduplicated_entries.append(entry)
                    else:
                        add_to_results("duplicated", entry.get("ID"), selection_result.results)
                else:
                    deduplicated_entries.append(entry)
            else:
                add_to_results("duplicated", entry.get("ID"), selection_result.results)
        else:
            deduplicated_entries.append(entry)

    db = BibDatabase()
    db.entries = deduplicated_entries
    writer = BibTexWriter()
    bibtex_str = writer.write(db)

    with open(file, 'w') as bibfile:
        bibfile.write(bibtex_str)

def get_quality_codes(citekey, selection_result):
    """Get the quality codes of a paper by its citekey"""
    quality_codes = []
    for n in range(1,11):
        qc = "qc" + str(n)
        if qc in selection_result.results and citekey in selection_result.results[qc]:
            quality_codes.append(qc.upper())

    return quality_codes

def is_study_included(entry, it_conf, selection_result):
    """Checks if a study meets the inclusion criteria and the required quality score"""
    citekey = entry.get("ID")
    if not citekey:
        return False
    
    for n in range(1, 11):
        ic_key = f"ic{n}"
        if ic_key in selection_result.results and citekey in selection_result.results[ic_key]:
            quality_score = get_quality_score(get_quality_codes(citekey, selection_result))
            return quality_score >= it_conf.threshold
    return False

def get_quality_score(quality_codes):
    """Get the quality score based on a paper's quality codes"""
    return len(quality_codes)

def add_to_results(key, citekey, results):
    """Add a paper to the set of results"""
    if not key or not citekey:
        raise ValueError("Key and citekey must be valid non-empty values")
    results.setdefault(key, []).append(citekey)

def get_venue(entry):
    """Get the venue of a paper"""
    if not entry or "ENTRYTYPE" not in entry:
        return ""
    
    entry_type = entry.get("ENTRYTYPE", '').lower()
    return {
        "inproceedings": entry.get("booktitle", ""),
        "book": entry.get("publisher",""),
        "journal": entry.get("article", "")
    }.get(entry_type, "")