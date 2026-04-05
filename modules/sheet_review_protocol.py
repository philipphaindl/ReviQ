import openpyxl
from openpyxl.styles import Font, Alignment
from modules.common import *
import os

def prepare_review_protocol(wb):
    """Create the selection protocol in Excel"""
    if not wb:
        raise ValueError("Workbook object cannot be None")
    
    if REVIEW_PROTOCOL_SHEET_NAME in wb.sheetnames:
        sheet = wb[REVIEW_PROTOCOL_SHEET_NAME]
    else:
        sheet = wb.create_sheet(REVIEW_PROTOCOL_SHEET_NAME)
    
    headers = [
        ("Date", 15),
        ("Iteration", 20),
        ("Library", 10),
        ("DOI", 40),
        ("Citekey", 40),
        ("Authors", 40),
        ("Title", 50),
        ("Year", 10, 'center'),
        ("Evaluation Method", 12, 'center', True),
        ("Duplicate", 10, 'center'),
        ("Exclusion Criteria", 12, 'center', True),
        ("Inclusion Criteria", 12, 'center', True),
        ("Quality Criteria", 12, 'center', True)
    ]
    
    for col_idx, (header, width, *alignment) in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        sheet.column_dimensions[cell.column_letter].width = width
        
        if alignment:
            cell.alignment = Alignment(horizontal=alignment[0], wrapText=True if len(alignment) > 1 else False)   

    return sheet

def update_review_protocol(wb, row_id, conf, selection_result):
    """Updates the selection protocol with the results of the given iteration"""
    process_manual_checks(conf, selection_result)
    sheet = wb[REVIEW_PROTOCOL_SHEET_NAME]
    while True:
        citekey = sheet.cell(row=row_id, column=5).value
        if citekey is None:
            break
        
        if "duplicated" in selection_result.results and citekey in selection_result.results["duplicated"]:
            sheet.cell(row=row_id, column=10, value="X").alignment = Alignment(horizontal='center', vertical='top')
        
        process_criteria(sheet, row_id, citekey, "ec", 11, "manual", conf.iteration_config, selection_result)
        process_criteria(sheet, row_id, citekey, "ic", 12, "manual", conf.iteration_config, selection_result)
        process_quality_criteria(sheet, row_id, citekey, selection_result)
        
        row_id += 1

def process_criteria(sheet, row_id, citekey, prefix, column, manual_label, it_conf, selection_result):
    """Processes and writes exclusion/inclusion criteria to the sheet."""
    for n in range(10):
        criteria = f"{prefix}{n}"
        if not sheet.cell(row=row_id, column=column).value:
            sheet.cell(row=row_id, column=column, value = "-").alignment = Alignment(horizontal='center', vertical='top')
        if criteria in selection_result.results and citekey in selection_result.results[criteria]:
            sheet.cell(row=row_id, column=column, value=criteria.upper()).alignment = Alignment(horizontal='center', vertical='top')
            if prefix == "ic":
                sheet.cell(row=row_id, column=9, value=manual_label)
            elif prefix == "ec" and n > 3: 
                sheet.cell(row=row_id, column=9, value=manual_label)

def process_quality_criteria(sheet, row_id, citekey, selection_result):
    """Processes and writes quality criteria to the sheet."""
    for n in range(10):
        qc = f"qc{n}"
        if not sheet.cell(row=row_id, column=13).value:
            sheet.cell(row=row_id, column=13, value = "-").alignment = Alignment(horizontal='center', vertical='top')
        if qc in selection_result.results and citekey in selection_result.results[qc]:
            existing_value = sheet.cell(row=row_id, column=13).value or ""
            new_value = ",".join(filter(None, [existing_value, qc.upper()]))
            if new_value.startswith("-"):
                new_value = new_value[2:]
            sheet.cell(row=row_id, column=13, value=new_value).alignment = Alignment(horizontal='center', vertical='top', wrapText=True)

def process_manual_checks(it_config, selection_result):
    """Processes manual checks from the Excel file and updates results and categories."""
    manual_check_excel = os.path.join(it_config.iteration_dir, AFTER_MANUAL_EVALUATION_DIRNAME, MANUAL_EVALUATION_FILENAME)
    if not os.path.exists(manual_check_excel):
        print(f"Error: Need file {manual_check_excel}")
        exit(1)
    
    wb = openpyxl.load_workbook(manual_check_excel)
    manual_check_sheet = wb.active
    row_id = 2
    
    while True:
        title = manual_check_sheet.cell(row=row_id, column=1).value
        if title is None:
            break
        
        title = title[:-4]
        author = manual_check_sheet.cell(row=row_id, column=2).value
        year = manual_check_sheet.cell(row=row_id, column=3).value
        citekey = get_citekey(title, author, year, selection_result.papers)
        
        if not citekey:
            print(f"Error: Unable to retrieve citekey for publication title: '{title}', author: {author}, year: {year} not found")
        
        update_results(manual_check_sheet, row_id, citekey, selection_result.results)
        update_categories(manual_check_sheet, row_id, citekey, selection_result.categories)
        
        row_id += 1

def update_results(sheet, row_id, citekey, results):
    """Updates the results dictionary based on exclusion and inclusion criteria."""
    for n in range(4, 7):
        val = sheet.cell(row=row_id, column=n).value
        if val:
            for single_val in val.lower().split(";"):
                if single_val not in results:
                    results[single_val] = []
                results[single_val].append(citekey)

def update_categories(sheet, row_id, citekey, categories):
    """Updates the categories dictionary with values from the manual check sheet."""
    column_id = 7
    while sheet.cell(row=row_id, column=column_id).value:
        column_name = sheet.cell(row=1, column=column_id).value
        column_value = sheet.cell(row=row_id, column=column_id).value
        
        if column_name not in categories:
            categories[column_name] = {}
        if column_value not in categories[column_name]:
            categories[column_name][column_value] = []
        
        categories[column_name][column_value].append(citekey)
        column_id += 1