from datetime import date
from openpyxl.styles import Font, Alignment
from modules.bibtex_processing import *
from openpyxl.utils import get_column_letter
from modules.common import *
from modules.selection import *

def process_selected_studies(wb, it_conf, selection_result):
    """Process and document the finally selected studies"""
    create_selected_studies_sheet(wb)
    fill_selected_studies_sheet(wb, it_conf, selection_result)

def create_selected_studies_sheet(wb):
    """Create Excel sheet documenting the selected studies"""
    wb.create_sheet(SELECTED_STUDIES_SHEET_NAME)
    selected_studies_sheet = wb[SELECTED_STUDIES_SHEET_NAME]

    selected_studies_sheet.cell(row=1, column=1).value = "Date"
    selected_studies_sheet.cell(row=1, column=1).font = Font(bold=True)
    selected_studies_sheet.column_dimensions["A"].width = 15

    selected_studies_sheet.cell(row=1, column=2).value = "Library"
    selected_studies_sheet.cell(row=1, column=2).font = Font(bold=True)
    selected_studies_sheet.column_dimensions["B"].width = 10

    selected_studies_sheet.cell(row=1, column=3).value = "Study Index"
    selected_studies_sheet.cell(row=1, column=3).font = Font(bold=True)
    selected_studies_sheet.column_dimensions["C"].width = 12

    selected_studies_sheet.cell(row=1, column=4).value = "Citekey"
    selected_studies_sheet.cell(row=1, column=4).font = Font(bold=True)
    selected_studies_sheet.column_dimensions["D"].width = 40

    selected_studies_sheet.cell(row=1, column=5).value = "Title"
    selected_studies_sheet.cell(row=1, column=5).font = Font(bold=True)
    selected_studies_sheet.column_dimensions["E"].width = 40

    selected_studies_sheet.cell(row=1, column=6).value = "Authors"
    selected_studies_sheet.cell(row=1, column=6).font = Font(bold=True)
    selected_studies_sheet.column_dimensions["F"].width = 40

    selected_studies_sheet.cell(row=1, column=7).value = "Venue"
    selected_studies_sheet.cell(row=1, column=7).font = Font(bold=True)
    selected_studies_sheet.column_dimensions["G"].width = 40

    selected_studies_sheet.cell(row=1, column=8).value = "Year"
    selected_studies_sheet.cell(row=1, column=8).font = Font(bold=True)
    selected_studies_sheet.cell(row=1, column=8).alignment = Alignment(horizontal="center")
    selected_studies_sheet.column_dimensions["H"].width = 10

    selected_studies_sheet.cell(row=1, column=9).value = "Type"
    selected_studies_sheet.cell(row=1, column=9).font = Font(bold=True)
    selected_studies_sheet.column_dimensions["I"].width = 20

    selected_studies_sheet.cell(row=1, column=10).value = "Quality Criteria"
    selected_studies_sheet.cell(row=1, column=10).font = Font(bold=True)
    selected_studies_sheet.column_dimensions["J"].width = 20

    selected_studies_sheet.cell(row=1, column=11).value = "Quality Score"
    selected_studies_sheet.cell(row=1, column=11).font = Font(bold=True)
    selected_studies_sheet.cell(row=1, column=11).alignment = Alignment(vertical='top', horizontal="center", wrapText=True)
    selected_studies_sheet.column_dimensions["K"].width = 15

    return selected_studies_sheet

def fill_selected_studies_sheet(wb, it_conf, selection_result):
    """Create the Excel sheet documenting the finally selected studies"""
    selected_studies_sheet = wb[SELECTED_STUDIES_SHEET_NAME]

    column_id = 12
    for column_name in selection_result.categories:
        selected_studies_sheet.cell(row=1, column=column_id).value = column_name
        selected_studies_sheet.cell(row=1, column=column_id).font = Font(bold=True)
        selected_studies_sheet.column_dimensions[get_column_letter(column_id)].width = 20
        column_id = column_id + 1

    directory = os.path.join(it_conf.iteration_dir, AFTER_MANUAL_EVALUATION_DIRNAME)

    resulting_entries = []
    study_index = 1
    for filename in os.listdir(directory):
        if filename.endswith('.bib'):
            final_entries = []
            with open(os.path.join(directory, filename)) as bibtex_file:
                parser = BibTexParser()
                parser.customization = convert_to_unicode  
                bib_database = bibtexparser.load(bibtex_file, parser=parser)
                for entry in bib_database.entries:
                    if is_study_included(entry, it_conf, selection_result):
                        resulting_entries.append(entry)
                        entry["note"] = "S" + str(study_index)

                        row_id = 1 + study_index
                        report_selected_study(selected_studies_sheet, filename, row_id, entry, selection_result)
                        study_index = study_index + 1

    db = BibDatabase()
    db.entries = resulting_entries

    writer = BibTexWriter()
    bibtex_str = writer.write(db)

    with open(FINAL_STUDIES_BIBFILE, 'w') as bibfile:
        bibfile.write(bibtex_str)

def report_selected_study(selected_studies_sheet, filename, row_id, entry, selection_result):
    """Report the details of a selected study in the Excel sheet"""
    selected_studies_sheet.cell(row=row_id, column=1).value = date.today()
    selected_studies_sheet.cell(row=row_id, column=1).alignment = Alignment(vertical='top')

    selected_studies_sheet.cell(row=row_id, column=2).value = filename[:-4]
    selected_studies_sheet.cell(row=row_id, column=2).alignment = Alignment(vertical='top')

    selected_studies_sheet.cell(row=row_id, column=3).value = entry.get("note")
    selected_studies_sheet.cell(row=row_id, column=3).alignment = Alignment(horizontal='center', vertical='top')

    citekey = entry.get("ID")
    selected_studies_sheet.cell(row=row_id, column=4).value = citekey
    selected_studies_sheet.cell(row=row_id, column=4).alignment = Alignment(vertical='top')

    selected_studies_sheet.cell(row=row_id, column=5).value = entry.get("title")
    selected_studies_sheet.cell(row=row_id, column=5).alignment = Alignment(vertical='top', wrapText=True)

    selected_studies_sheet.cell(row=row_id, column=6).value = entry.get("author")
    selected_studies_sheet.cell(row=row_id, column=6).alignment = Alignment(vertical='top', wrapText=True)

    entry_type = entry.get("ENTRYTYPE", '')
    if "article" == entry_type:
        selected_studies_sheet.cell(row=row_id, column=7).value = entry.get("journal")
        selected_studies_sheet.cell(row=row_id, column=9).value = "Journal Article"
    elif "inproceedings" == entry_type:
        selected_studies_sheet.cell(row=row_id, column=7).value = entry.get("booktitle")
        selected_studies_sheet.cell(row=row_id, column=9).value = "Conference Paper"
    elif "book" == entry_type:
        selected_studies_sheet.cell(row=row_id, column=7).value = entry.get("publisher")
        selected_studies_sheet.cell(row=row_id, column=9).value = "Book"
    else:
        selected_studies_sheet.cell(row=row_id, column=7).value = "n/a"
        selected_studies_sheet.cell(row=row_id, column=9).value = "n/a"

    selected_studies_sheet.cell(row=row_id, column=7).alignment = Alignment(vertical='top', wrapText=True)
    selected_studies_sheet.cell(row=row_id, column=9).alignment = Alignment(vertical='top', wrapText=True)

    selected_studies_sheet.cell(row=row_id, column=8).value = entry.get("year")
    selected_studies_sheet.cell(row=row_id, column=8).alignment = Alignment(vertical='top', horizontal="center")

    quality_codes = get_quality_codes(citekey, selection_result)
    quality_score = get_quality_score(quality_codes)

    selected_studies_sheet.cell(row=row_id, column=10).value = ",".join(quality_codes)
    selected_studies_sheet.cell(row=row_id, column=10).alignment = Alignment(vertical='top')

    selected_studies_sheet.cell(row=row_id, column=11).value = str(quality_score)
    selected_studies_sheet.cell(row=row_id, column=11).alignment = Alignment(vertical='top', horizontal="center")

    column_id = 12
    study_category_label = ""
    for column_name in selection_result.categories:
        study_category_label = ""
        for category_label in selection_result.categories[column_name]:
            if citekey in selection_result.categories[column_name][category_label]:
                study_category_label = category_label
                break

        selected_studies_sheet.cell(row=row_id, column=column_id).value = study_category_label
        selected_studies_sheet.cell(row=row_id, column=column_id).alignment = Alignment(vertical='top')
        column_id = column_id + 1

def write_selected_studies(it_config, selection_result):
    """Processes and writes the final bibliography files after manual checks."""
    directory = os.path.join(it_config.iteration_dir, NEEDING_MANUAL_EVALUATION_DIRNAME)
    total_selected = 0
    
    for filename in os.listdir(directory):
        if filename.endswith('.bib'):
            final_entries = []
            with open(os.path.join(directory, filename)) as bibtex_file:
                parser = BibTexParser()
                bib_database = parser.parse_file(bibtex_file)
                
                for entry in bib_database.entries:
                    citekey = entry.get("ID")
                    if is_study_included(entry, it_config, selection_result):
                        final_entries.append(entry)
                        total_selected += 1
            
            write_bibtex_entries(os.path.join(it_config.iteration_dir, AFTER_MANUAL_EVALUATION_DIRNAME, filename), final_entries)
    
    return total_selected