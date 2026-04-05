import openpyxl
import sys
import os
from modules.common import *

def write_to_excel(papers, code_names, dest):
	wb = openpyxl.Workbook()
	sheet = wb.active

	sheet.cell(row=1, column=1).value = "Title"
	sheet.cell(row=1, column=2).value = "Author"
	sheet.cell(row=1, column=3).value = "Year"

	col_index = 4
	for code_name in code_names:
		sheet.cell(row=1, column=col_index).value = code_name
		col_index = col_index + 1

	row = 2
	for paper_key in papers:
		paper = papers[paper_key]
		sheet.cell(row=row, column=1).value = paper.title
		sheet.cell(row=row, column=2).value = paper.author
		sheet.cell(row=row, column=3).value = paper.year

		col_index = 4
		for code_name in code_names:
			code_value_concatenated = ""
			if code_name in paper.codes:

				for code_value in paper.codes[code_name]:
					code_value_concatenated = code_value_concatenated + ";" + code_value
				code_value_concatenated = code_value_concatenated[1:]

			sheet.cell(row=row, column=col_index).value = code_value_concatenated
			col_index = col_index + 1

		row = row + 1

	wb.save(dest)

if len(sys.argv) < 2:
	print("Usage: python <iteration_dir> <sourcefile_name>")
	exit(1)

src = os.path.join(sys.argv[1], sys.argv[2])
dest = os.path.join(sys.argv[1], "manual_evaluations.xlsx")
papers = {}

wb_obj = openpyxl.load_workbook(src)
sheet_obj = wb_obj.active

field_name = sheet_obj.cell(row=1, column=4)
if field_name.value in ["Dokumentname", "Document Name"]:
	
	row = 2
	doc_name = field_name.value
	while True:
		doc_name = sheet_obj.cell(row=row, column=4).value
		code = sheet_obj.cell(row=row, column=5).value
		row = row + 1
		if doc_name != None:
			doc_parts = doc_name.split(" - ")
			if len(doc_parts) >= 3:
				author = doc_parts[0].strip()
				year = doc_parts[1].strip()
				title = doc_parts[2].strip()
				p = None
				if doc_name in papers:
					p = papers.get(doc_name)
				else:
					p = Paper(author, title, year)
					papers[doc_name] = p

				code_parts = code.split(">")
				if code.startswith("Inclusion") or code.startswith("Exclusion") or code.startswith("Quality") or code.startswith("Research Types") or code.startswith("Contribution Types"):
					code_name = code_parts[0].strip()
					criterion = code_parts[1].strip()

					p.append_code(code_name, criterion)
		else:
			break

	code_names = set()
	for key in papers:
		p = papers[key]
		for code in p.codes:
			code_names.add(code)

	write_to_excel(papers, code_names, dest)

else:
	print("Error finding field starting point for conversion (column with document name)")

